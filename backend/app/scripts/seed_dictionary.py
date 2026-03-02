"""Seed the database from the Excel dictionary file."""

import logging
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.word import Word, RootFamily, RootFamilyMember

logger = logging.getLogger(__name__)

EXCEL_PATH = Path("/app/data/dictionary.xlsx")

FREQUENCY_MAP = {
    "high": 1,
    "mid": 2,
    "low": 3,
    "rare": 4,
}

# Map Russian binyan names to DB IDs (seeded in migration 001)
BINYAN_MAP = {
    "паъаль": 1, "пааль": 1,
    "нифъаль": 2,
    "пиъэль": 3, "пиэль": 3,
    "пуъаль": 4, "пуаль": 4,
    "hифъиль": 5, "хифъиль": 5, "гифъиль": 5,
    "hуфъаль": 6, "хуфъаль": 6, "гуфъаль": 6,
    "hитпаъэль": 7, "хитпаэль": 7, "гитпаъэль": 7,
}

POS_MAP = {
    "сущ": "noun",
    "сущ.": "noun",
    "гл": "verb",
    "гл.": "verb",
    "прил": "adj",
    "прил.": "adj",
    "нар": "adv",
    "нар.": "adv",
    "предл": "prep",
    "предл.": "prep",
    "союз": "conj",
    "мест": "pron",
    "мест.": "pron",
    "числ": "num",
    "числ.": "num",
    "част": "particle",
    "част.": "particle",
    "межд": "interj",
    "межд.": "interj",
}


def _normalize_pos(raw: str | None) -> str | None:
    if not raw:
        return None
    key = raw.strip().lower().rstrip(".")
    # Try with and without dot
    return POS_MAP.get(key, POS_MAP.get(key + ".", raw.strip()))


def _normalize_frequency(raw: str | None) -> int | None:
    if not raw:
        return None
    return FREQUENCY_MAP.get(raw.strip().lower())


def _normalize_binyan(raw: str | None) -> str | None:
    """Return the normalized binyan name (or None)."""
    if not raw:
        return None
    return raw.strip().lower()


async def seed_dictionary(db: AsyncSession) -> dict:
    """Parse the Excel dictionary and insert words + root families.

    Returns a summary dict with counts.
    """
    count = await db.scalar(select(func.count()).select_from(Word))
    if count and count > 0:
        logger.info("Dictionary already seeded (%d words). Skipping.", count)
        return {"words_existing": count, "words_added": 0, "roots_added": 0}

    if not EXCEL_PATH.exists():
        logger.warning("Dictionary Excel file not found at %s", EXCEL_PATH)
        return {"error": "Excel file not found"}

    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    # --- Sheet 1: Main dictionary ---
    ws = wb["Весь словарь Бет"]
    words_added = 0
    word_map: dict[str, int] = {}  # hebrew -> word.id

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if not row or not row[0]:
            continue

        hebrew = str(row[0]).strip()
        transliteration = str(row[1]).strip() if row[1] else None
        translation_ru = str(row[2]).strip() if row[2] else ""
        pos_raw = str(row[3]).strip() if row[3] else None
        binyan_raw = str(row[4]).strip() if row[4] else None
        root = str(row[5]).strip() if row[5] else None
        # row[6] = related words (we'll use root families instead)
        frequency_raw = str(row[7]).strip() if row[7] else None
        # row[8] = source

        if not translation_ru:
            continue

        pos = _normalize_pos(pos_raw)
        frequency_rank = _normalize_frequency(frequency_raw)

        word = Word(
            hebrew=hebrew,
            transliteration=transliteration,
            translation_ru=translation_ru,
            pos=pos,
            root=root,
            frequency_rank=frequency_rank,
            level_id=2,  # Bet level
        )

        # Store binyan info in the pos field for verbs
        if binyan_raw and pos == "verb":
            normalized_binyan = _normalize_binyan(binyan_raw)
            # We'll store original binyan name in transliteration metadata
            # The binyan_id link is through verb_conjugations table

        db.add(word)
        await db.flush()
        word_map[hebrew] = word.id
        words_added += 1

    # --- Sheet 2: Root families ---
    ws2 = wb["Семьи корней"]
    roots_added = 0
    root_cache: dict[str, int] = {}  # root string -> root_family.id

    rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
    for row in rows2:
        if not row or not row[0]:
            continue

        root_str = str(row[0]).strip()
        word_hebrew = str(row[1]).strip() if row[1] else None

        if not root_str or not word_hebrew:
            continue

        # Create root family if not exists
        if root_str not in root_cache:
            family = RootFamily(root=root_str)
            db.add(family)
            await db.flush()
            root_cache[root_str] = family.id
            roots_added += 1

        # Link word to root family
        word_id = word_map.get(word_hebrew)
        if word_id:
            member = RootFamilyMember(
                root_family_id=root_cache[root_str],
                word_id=word_id,
            )
            db.add(member)

    wb.close()
    await db.commit()

    logger.info("Seeded %d words and %d root families.", words_added, roots_added)
    return {"words_added": words_added, "roots_added": roots_added}
