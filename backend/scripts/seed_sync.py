"""Synchronous dictionary seeder — runs between Alembic migration stages."""
import os
import uuid
from pathlib import Path

from passlib.context import CryptContext
from openpyxl import load_workbook
from sqlalchemy import create_engine, text

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

EXCEL_PATH = Path("/app/data/dictionary.xlsx")

POS_MAP = {
    "сущ": "noun", "сущ.": "noun",
    "гл": "verb", "гл.": "verb",
    "прил": "adj", "прил.": "adj",
    "нар": "adv", "нар.": "adv",
    "предл": "prep", "предл.": "prep",
    "союз": "conj",
    "мест": "pron", "мест.": "pron",
    "числ": "num", "числ.": "num",
    "част": "particle", "част.": "particle",
    "межд": "interj", "межд.": "interj",
}

FREQUENCY_MAP = {"high": 1, "mid": 2, "low": 3, "rare": 4}


def _pos(raw):
    if not raw:
        return None
    key = raw.strip().lower().rstrip(".")
    return POS_MAP.get(key, POS_MAP.get(key + ".", raw.strip()))


def main():
    db_url = os.environ.get("DATABASE_URL", "")
    sync_url = db_url.replace("+asyncpg", "").replace("postgresql://", "postgresql+psycopg2://")
    if "psycopg2" not in sync_url:
        sync_url = sync_url.replace("postgresql://", "postgresql+psycopg2://")

    engine = create_engine(sync_url)
    with engine.connect() as conn:
        count = conn.execute(text("SELECT COUNT(*) FROM words")).scalar()
        if count and count > 0:
            print(f"Dictionary already seeded ({count} words). Skipping.")
            return

        if not EXCEL_PATH.exists():
            print(f"Excel file not found at {EXCEL_PATH}. Skipping seed.")
            return

        wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb["Весь словарь Бет"]
        added = 0
        word_map = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            hebrew = str(row[0]).strip()
            transliteration = str(row[1]).strip() if row[1] else None
            translation_ru = str(row[2]).strip() if row[2] else ""
            pos = _pos(str(row[3]).strip() if row[3] else None)
            root = str(row[5]).strip() if row[5] else None
            freq_raw = str(row[7]).strip().lower() if row[7] else None
            frequency_rank = FREQUENCY_MAP.get(freq_raw) if freq_raw else None

            if not translation_ru:
                continue

            result = conn.execute(
                text(
                    "INSERT INTO words (hebrew, transliteration, translation_ru, pos, root, frequency_rank, level_id) "
                    "VALUES (:h, :tr, :tru, :p, :r, :f, 2) RETURNING id"
                ),
                {"h": hebrew, "tr": transliteration, "tru": translation_ru, "p": pos, "r": root, "f": frequency_rank},
            )
            wid = result.scalar()
            word_map[hebrew] = wid
            added += 1

        # Root families
        ws2 = wb["Семьи корней"]
        root_cache = {}
        roots_added = 0

        for row in ws2.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            root_str = str(row[0]).strip()
            word_hebrew = str(row[1]).strip() if row[1] else None
            if not root_str or not word_hebrew:
                continue

            if root_str not in root_cache:
                rid = conn.execute(
                    text("INSERT INTO root_families (root) VALUES (:r) RETURNING id"),
                    {"r": root_str},
                ).scalar()
                root_cache[root_str] = rid
                roots_added += 1

            wid = word_map.get(word_hebrew)
            if wid:
                conn.execute(
                    text("INSERT INTO root_family_members (root_family_id, word_id) VALUES (:rf, :w)"),
                    {"rf": root_cache[root_str], "w": wid},
                )

        wb.close()

        # Seed default user
        exists = conn.execute(text("SELECT 1 FROM users WHERE email = :e"), {"e": "admin@ulpan.ai"}).first()
        if not exists:
            uid = str(uuid.uuid4())
            pw_hash = pwd_context.hash("admin123")
            conn.execute(
                text(
                    "INSERT INTO users (id, email, password_hash, display_name, native_lang, current_level, xp, streak_days) "
                    "VALUES (:id, :e, :pw, :n, 'ru', 1, 0, 0)"
                ),
                {"id": uid, "e": "admin@ulpan.ai", "pw": pw_hash, "n": "Admin"},
            )
            print("Seeded default user: admin@ulpan.ai / admin123")

        conn.commit()
        print(f"Seeded {added} words, {roots_added} root families.")


if __name__ == "__main__":
    main()
